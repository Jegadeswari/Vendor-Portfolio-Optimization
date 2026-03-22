#!/usr/bin/env python3
"""
restore-formatting.py
---------------------
Copies all formatting from the original template onto the output workbook
produced by update-vendor-analysis-sheet.js, preserving the data values
already written (Department, Description, Suggestion).

Usage:
  python3 scripts/restore-formatting.py <template.xlsx> <output.xlsx>

The output file is updated in-place.
"""

import sys
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

template_path = sys.argv[1]
output_path   = sys.argv[2]

print(f"Template : {template_path}")
print(f"Output   : {output_path}")

wb_tmpl = openpyxl.load_workbook(template_path)
wb_out  = openpyxl.load_workbook(output_path)

# ── 1. Vendor Analysis Assessment — restore formatting on all cells ──────────

ws_tmpl = wb_tmpl['Vendor Analysis Assessment']
ws_out  = wb_out['Vendor Analysis Assessment']

for row in ws_tmpl.iter_rows():
    for cell_tmpl in row:
        cell_out = ws_out.cell(row=cell_tmpl.row, column=cell_tmpl.column)

        # Always copy formatting
        if cell_tmpl.font:       cell_out.font       = copy(cell_tmpl.font)
        if cell_tmpl.fill:       cell_out.fill       = copy(cell_tmpl.fill)
        if cell_tmpl.border:     cell_out.border     = copy(cell_tmpl.border)
        if cell_tmpl.alignment:  cell_out.alignment  = copy(cell_tmpl.alignment)
        if cell_tmpl.number_format: cell_out.number_format = cell_tmpl.number_format

        # Only restore VALUES on Vendor Analysis Assessment
        # for columns the pipeline did not write.
        # All other sheets keep the values written by the pipeline steps.
        if ws_tmpl.title == 'Vendor Analysis Assessment':
            col = cell_tmpl.column
            if col not in (2, 4, 5):  # skip B (dept), D (desc), E (suggestion)
                cell_out.value = cell_tmpl.value
        # For Top 3 Opportunities, Methodology, CEOCFO Recommendations:
        # do NOT restore values — keep what the pipeline wrote

# ── 2. Restore column widths ─────────────────────────────────────────────────

for col_letter, dim in ws_tmpl.column_dimensions.items():
    ws_out.column_dimensions[col_letter].width = dim.width

# ── 3. Restore row heights ───────────────────────────────────────────────────

for row_num, dim in ws_tmpl.row_dimensions.items():
    ws_out.row_dimensions[row_num].height = dim.height

# ── 4. Restore freeze panes ──────────────────────────────────────────────────

ws_out.freeze_panes = ws_tmpl.freeze_panes

# ── 5. Restore all other sheets from template that are missing in output ─────

for sheet_name in wb_tmpl.sheetnames:
    if sheet_name == 'Vendor Analysis Assessment':
        continue  # already handled above
    if sheet_name not in wb_out.sheetnames:
        print(f"  Restoring missing sheet: {sheet_name}")
        ws_src = wb_tmpl[sheet_name]
        ws_dst = wb_out.create_sheet(sheet_name)
        for row in ws_src.iter_rows():
            for cell in row:
                new_cell = ws_dst.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.font:       new_cell.font       = copy(cell.font)
                if cell.fill:       new_cell.fill       = copy(cell.fill)
                if cell.border:     new_cell.border     = copy(cell.border)
                if cell.alignment:  new_cell.alignment  = copy(cell.alignment)
                if cell.number_format: new_cell.number_format = cell.number_format
        for col_letter, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col_letter].width = dim.width
        for row_num, dim in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[row_num].height = dim.height
        ws_dst.freeze_panes = ws_src.freeze_panes

# ── 6. Restore formatting on output sheets that already exist ───────────────
# These sheets were written by pipeline steps 09, 11, 12.
# Restore their formatting from template but preserve their VALUES.

output_only_sheets = ['Top 3 Opportunities', 'Methodology', 'CEOCFO Recommendations']

for sheet_name in output_only_sheets:
    if sheet_name not in wb_tmpl.sheetnames:
        continue
    if sheet_name not in wb_out.sheetnames:
        continue  # handled by section 5 above

    ws_src = wb_tmpl[sheet_name]
    ws_dst = wb_out[sheet_name]

    for row in ws_src.iter_rows():
        for cell in row:
            cell_out = ws_dst.cell(row=cell.row, column=cell.column)
            # Restore formatting only — never overwrite pipeline-written values
            if cell.font:          cell_out.font          = copy(cell.font)
            if cell.fill:          cell_out.fill          = copy(cell.fill)
            if cell.border:        cell_out.border        = copy(cell.border)
            if cell.alignment:     cell_out.alignment     = copy(cell.alignment)
            if cell.number_format: cell_out.number_format = cell.number_format

    for col_letter, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = dim.width
    for row_num, dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_num].height = dim.height
    ws_dst.freeze_panes = ws_src.freeze_panes
    print(f"  Formatting restored for sheet: {sheet_name}")
    
# ── 7. Set wrap text on all cells and fit columns to content ───────────────

def apply_wrap_and_fit(ws):
    """Set wrap_text=True on all cells and set column width to fit content."""
    col_widths = {}

    for row in ws.iter_rows():
        for cell in row:
            # Set wrap text
            existing = cell.alignment
            cell.alignment = Alignment(
                wrap_text      = True,
                horizontal     = existing.horizontal     if existing else None,
                vertical       = existing.vertical       if existing else None,
                text_rotation  = existing.text_rotation  if existing else 0,
                shrink_to_fit  = existing.shrink_to_fit  if existing else False,
                indent         = existing.indent         if existing else 0
            )
            # Track max content length per column
            if cell.value:
                col_letter = get_column_letter(cell.column)
                content_len = len(str(cell.value).split('\n')[0])  # first line only
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), content_len)

    # Apply column widths — cap at 60 chars to avoid excessively wide columns
    for col_letter, max_len in col_widths.items():
        fitted_width = min(max_len + 2, 60)  # +2 padding, 60 char cap
        ws.column_dimensions[col_letter].width = fitted_width
        ws.column_dimensions[col_letter].bestFit = True

# Apply to all sheets in the output workbook
for sheet_name in wb_out.sheetnames:
    apply_wrap_and_fit(wb_out[sheet_name])
    print(f"  Wrap + fit applied to sheet: {sheet_name}")

# ── 8. Preserve sheet order from template ────────────────────────────────────

template_order = wb_tmpl.sheetnames
for i, name in enumerate(template_order):
    if name in wb_out.sheetnames:
        wb_out.move_sheet(name, offset=i - wb_out.sheetnames.index(name))

# ── 9. Save ───────────────────────────────────────────────────────────────────

wb_out.save(output_path)
print(f"Formatting restored and saved to: {output_path}")